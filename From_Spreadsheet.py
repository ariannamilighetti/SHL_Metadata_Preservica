import metadata_creator
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import os
import openpyxl
import webbrowser

class From_Spreadsheet (ttk.Frame):

    def __init__(self, master):
        super().__init__(master)
        self.selection_frm = ttk.LabelFrame(self, text="Items")
        self.selection_frm.grid(row=0, column=0, sticky='nwse', padx=5, pady=5)
        self.spreadsheeet_field()
        self.folder_fields()
        self.run_button()
        self.columnconfigure(0, weight = 1)
        self.rowconfigure(0, weight =0)
        self.selection_frm.columnconfigure(0, weight = 1)
        self.selection_frm.columnconfigure(1, weight = 0)
        self.selection_frm.columnconfigure(2, weight = 2)
    
    def callback(self, url):
        webbrowser.open_new_tab(url)

    def spreadsheeet_field(self):
        # Spreadsheet Fields
        template_label = tk.Label(master=self.selection_frm, text="A template of the correct metadata spreadsheet can be found here:")
        link = tk.Label(self.selection_frm, text="SHL Digitisation Sharepoint > Preservica Prep > Metadata Template for Preservica", fg='blue', cursor="hand2")
        template_label.grid(column=0, row=0, columnspan=3, sticky="nw", padx=5)
        link.grid(column=0, row=1, columnspan=3, sticky="nw", padx=5)
        link.bind("<Button-1>", lambda e: self.callback("https://uolonline.sharepoint.com/:x:/s/shl/bd/EQqjji8IPAVKgX4y2Qp02fkBLAJq1CnNYHn_0BtrOydGaw?e=Zsh7EL"))
        spreadsheet_label = tk.Label(master=self.selection_frm, text="Import metadata spreadsheet", anchor='nw')
        spreadsheet_button = tk.Button(self.selection_frm, text='Open', command=self.upload_spreadsheet,bg='#ff8962')
        self.spreadsheet_filename = tk.Entry(master=self.selection_frm)
        spreadsheet_label.grid(column=0, row=2, columnspan=1, sticky="ew", padx=5, pady=5)
        spreadsheet_button.grid(column=1, row=2, columnspan=1, sticky="ew", padx=5, pady=5)
        self.spreadsheet_filename.grid(column=2, row=2, columnspan=3, sticky="ew", padx=5, pady=5)
    
    def folder_fields(self):
        # Folder Fields
        dir_label = tk.Label(master=self.selection_frm, text="Select item(s) parent folder", anchor='nw')
        dir_button = tk.Button(self.selection_frm, text='Open', command=self.askDirectory_ss, bg='#ff8962')
        self.ss_directory_label = tk.Entry(master=self.selection_frm)
        dir_label.grid(column=0, row=3, columnspan=1, sticky="ew", padx=5, pady=5)
        dir_button.grid(column=1, row=3, columnspan=1, sticky="ew", padx=5, pady=5)
        self.ss_directory_label.grid(column=2, row=3, columnspan=3, sticky="ew", padx=5, pady=5)

    def check_file_format(self, excel_file):
        correct_setup = ['Title', 'Classmark', 'Catalogue number', 'Barcode', 'SHL Collection', 'Digitised by', 'Digitisation Date', 'Access Conditions', 'Restriction reason', 'Restriction expiry date', 'Copyright status', 'Copyright Statement', 'Licence details', 'Number of Images', 'Digi date text',	'Restr Exp Date']
        validation = True
        i = 0
        for col in excel_file.iter_cols(4, excel_file.max_column):

            if col[0].value == correct_setup[i]:
                i+=1
                continue
            else :
                validation = False
                break
        return validation

    def read_metadata_file(self):
        openpyxl.reader.excel.warnings.simplefilter(action='ignore')
        dataframe = openpyxl.load_workbook(self.metadata_file, data_only=True)
        excel_file = dataframe.active
        metadata = []
        required_columns = [6, 3, 4, 5, 7, 8, 17, 10, 11, 18, 13, 14, 15]
        validation = self.check_file_format(excel_file)
        if validation:
            for row in range(1, excel_file.max_row+1):
                metadata.append([])
                for i in required_columns:
                    if excel_file.cell(row=row, column=i+1).value is None:
                        metadata[row-1].append('')
                    else: 
                        metadata[row-1].append(excel_file.cell(row=row, column=i+1).value)
        else :
            print("file format incorrect")

        return metadata

    def find_item_metadata(self):
        item_images = []
        for directory in os.listdir(self.image_folder):
            item_images.append(directory)
        excel_metadata = self.read_metadata_file()
        metadata = {}
        for item in excel_metadata:
            if str(item[0]) in str(item_images):
                metadata[item[0]] = item
        return metadata

    def validate_data(self):
        self.image_folder = self.ss_directory_label.get()
        self.metadata_file = self.spreadsheet_filename.get()
        self.metadata_dictionary = self.find_item_metadata()
        dict_keys = self.metadata_dictionary.keys()
        row = 0
        outcome_frame = tk.Frame(self)
        outcome_frame.pack(fill='both', expand=True)
        barcode_l = tk.Label(outcome_frame, text='Barcode')
        barcode_l.grid( row=row, column = 0, padx=10, pady=5)
        collection_l = tk.Label(outcome_frame, text='Collection')
        collection_l.grid(row=row, column = 1, padx=10, pady=5)
        digitised_by_l = tk.Label(outcome_frame, text='Digitised By')
        digitised_by_l.grid(row=row, column = 2, padx=10, pady=5)
        access_cond_l = tk.Label(outcome_frame, text='Access Conditions')
        access_cond_l.grid(row=row, column = 3, padx=10, pady=5)
        restr_reason_l = tk.Label(outcome_frame, text='Restriction Reason')
        restr_reason_l.grid(row=row, column = 4, padx=10, pady=5)
        cop_status_l = tk.Label(outcome_frame,  text='Copyright Status')
        cop_status_l.grid(row=row, column = 5, padx=10, pady=5)
        rights_stat_l = tk.Label(outcome_frame, text='Rights Statement')
        rights_stat_l.grid(row=row, column = 6, padx=10, pady=5)
        folder_l = tk.Label(outcome_frame, text='Folder')
        folder_l.grid(row=row, column = 7, padx=10, pady=5)
        succ_l = tk.Label(outcome_frame, text='Overall Success')
        succ_l.grid(row=row, column = 8, padx=10, pady=5)

        for key in dict_keys:
            row += 1
            column = 0
            mets = self.metadata_dictionary.get(key)
            folder = self.ss_directory_label.get() + "/" + str(key)
            labels = metadata_creator.create_metadata(folder, mets, self)
            barcode_label = tk.Label(outcome_frame, text=key)
            barcode_label.grid(row=row, column = column)
            for i in labels:
                column += 1
                label = tk.Label(outcome_frame,  text=i[0], wraplength=100, background=i[1])
                label.grid(row=row, column=column, padx=10, pady=5)

    def run_button(self):
        run_button = tk.Button(self, text="Run", command=lambda: self.validate_data(), bg='#7fd1ae')
        run_button.grid(row=2, column=0, sticky='nwse', padx=5)

    def upload_spreadsheet (self,event=None):
        global spreadsheet_input
        spreadsheet_input = filedialog.askopenfilename(filetypes=(
            ("Excel files", "*.xlsx"), ("Any file", "*")))
        self.spreadsheet_filename.delete(0, "end")
        self.spreadsheet_filename.insert(0, spreadsheet_input)

    def askDirectory_ss(self, event=None):
        global end_directory
        end_directory = filedialog.askdirectory()
        self.ss_directory_label.delete(0, "end")
        self.ss_directory_label.insert(0, end_directory)
  
    def complete_label(self, label) :
        complete_label = tk.Label(self, text=label)
        complete_label.grid(row=2, column=0, sticky='nwse', padx=5, pady=5)
 

# to test this class
def __main__():
    app = tk.Tk()
    tabs = ttk.Notebook(app)
    tabs.grid(row=1, column= 0, sticky="nsew")
    # Create Tabs
    data_input = From_Spreadsheet(master=tabs)
    tabs.add(data_input, text='From Input')
    tk.mainloop()

if __name__ == '__main__':
    __main__()